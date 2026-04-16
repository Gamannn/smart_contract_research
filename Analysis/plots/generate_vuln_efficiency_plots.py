"""
generate_vuln_efficiency_plots.py
==================================
Reads vulnerability_tool_efficiency_report_batchXX.xlsx files from
analysis/batch_0X/ folders and produces publication-quality PNG plots
for the IEEE smart-contract obfuscation research paper.

Output folder : analysis/plots/
Run from      : smart_contract_research/ root  OR  analysis/plots/
                (script auto-detects the base path)

Plots generated
---------------
 01_overall_efficiency_all_batches.png   - grouped bar: orig vs obf, all 5 batches
 02_per_vuln_orig_B01..B05.png           - per-vulnerability efficiency (original), one per batch
 03_per_vuln_obf_B01..B05.png            - per-vulnerability efficiency (obfuscated), one per batch
 04_tool_trend_orig_vs_obf.png           - line chart: overall efficiency trend across batches
 05_heatmap_orig_avg.png                 - heatmap: avg efficiency tool×vuln (original)
 06_heatmap_obf_avg.png                  - heatmap: avg efficiency tool×vuln (obfuscated)
 07_radar_orig_per_batch.png             - radar charts: per-vulnerability per-tool (original)
 08_nv_positives_stacked.png             - stacked bar: N(v) consensus positives per vuln×batch
 09_efficiency_drop_heatmap.png          - heatmap: efficiency drop (orig→obf) per tool×batch
 10_tool_robustness_bar.png              - bar chart: robustness score per tool per batch
"""

import os, sys, re
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.colors as mcolors
import warnings
warnings.filterwarnings('ignore')

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl --break-system-packages")

# ─────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Resolve base: script lives at analysis/plots/ → base = analysis/
BASE = os.path.abspath(os.path.join(SCRIPT_DIR, '..'))
OUT  = os.path.join(BASE, 'plots')
os.makedirs(OUT, exist_ok=True)

BATCHES      = ['B01', 'B02', 'B03', 'B04', 'B05']
BATCH_NUMS   = ['01',  '02',  '03',  '04',  '05']
TOOLS        = ['Mythril', 'Slither', 'Oyente', 'Osiris']
VULNS_FULL   = ['Reentrancy', 'Integer', 'Access', 'TOD', 'DoS',
                'Delegatecall', 'Unchecked Call', 'Timestamp', 'Frozen Ether']
VULNS_PLOT   = ['Reentrancy', 'Integer', 'Access', 'TOD', 'DoS',
                'Unchecked Call', 'Timestamp', 'Frozen Ether']   # skip Delegatecall (sparse)

COLORS  = {'Mythril': '#2E75B6', 'Slither': '#375623',
           'Oyente':  '#C55A11', 'Osiris':  '#7030A0'}
BCOLORS = ['#1F4E79', '#833C00', '#375623', '#843C0C', '#4472C4']

# ─────────────────────────────────────────────────────────────
# DATA INGESTION  — reads directly from xlsx
# ─────────────────────────────────────────────────────────────
def _to_float(v):
    """Convert cell value to float, returning None for N/A / empty."""
    if v is None:
        return None
    s = str(v).strip().replace('%', '').replace('N/A', '').replace('—', '').strip()
    if s == '' or s.lower() == 'n/a':
        return None
    try:
        return float(s)
    except ValueError:
        return None

def _find_header_row(ws):
    """Return 0-based index of the row whose first cell is 'Vulnerability'."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and str(row[0]).strip().lower() == 'vulnerability':
            return i
    return None

def load_batch(batch_num):
    """
    Returns dict: {
        'orig': {vuln: {tool: float|None}, ..., 'OVERALL': {tool: float|None}},
        'obf' : same,
        'nv_orig': {vuln: int},
        'nv_obf' : {vuln: int},
        'tool_comparison': {tool: {'orig': float, 'obf': float, 'drop': float, 'robustness': float}},
    }
    """
    path = os.path.join(BASE, f'batch_{batch_num}',
                        f'vulnerability_tool_efficiency_report_batch{batch_num}.xlsx')
    if not os.path.exists(path):
        print(f"  [WARN] File not found: {path}")
        return None

    wb = openpyxl.load_workbook(path, data_only=True)

    def _parse_eff_sheet(sheet_name):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        hi = _find_header_row(ws)
        if hi is None:
            return {}, {}
        data_rows = rows[hi+1:]
        eff   = {}
        nv    = {}
        for row in data_rows:
            if not row or row[0] is None:
                continue
            vuln = str(row[0]).strip()
            if not vuln or vuln.lower() in ('vulnerability', 'overall'):
                if str(row[0]).strip().upper() == 'OVERALL':
                    e = {}
                    for ti, tool in enumerate(TOOLS):
                        col = 3 + ti * 2   # Efficiency columns: 3,5,7,9
                        e[tool] = _to_float(row[col]) if col < len(row) else None
                    eff['OVERALL'] = e
                continue
            # canonical name
            canon = vuln
            for v in VULNS_FULL + ['OVERALL']:
                if v.lower() in vuln.lower() or vuln.lower() in v.lower():
                    canon = v
                    break
            n_pos = row[1] if len(row) > 1 else None
            try:
                nv[canon] = int(n_pos) if n_pos is not None else 0
            except (ValueError, TypeError):
                nv[canon] = 0
            e = {}
            for ti, tool in enumerate(TOOLS):
                col = 3 + ti * 2
                e[tool] = _to_float(row[col]) if col < len(row) else None
            eff[canon] = e
        return eff, nv

    orig_eff, nv_orig = _parse_eff_sheet('Original_Efficiency')
    obf_eff,  nv_obf  = _parse_eff_sheet('Obfuscated_Efficiency')

    # Tool_Comparison sheet
    tc = {}
    ws_tc = wb['Tool_Comparison']
    rows_tc = list(ws_tc.iter_rows(values_only=True))
    hi_tc = None
    for i, row in enumerate(rows_tc):
        if row and str(row[0]).strip().lower() == 'tool':
            hi_tc = i
            break
    if hi_tc is not None:
        for row in rows_tc[hi_tc+1:]:
            if not row or row[0] is None:
                continue
            tool = str(row[0]).strip()
            if tool not in TOOLS:
                continue
            tc[tool] = {
                'orig':       _to_float(row[1]),
                'obf':        _to_float(row[2]),
                'drop':       _to_float(row[3]),
                'robustness': _to_float(row[4]) if len(row) > 4 else None,
            }

    return {'orig': orig_eff, 'obf': obf_eff,
            'nv_orig': nv_orig, 'nv_obf': nv_obf,
            'tool_comparison': tc}

print("Loading data from xlsx files...")
DATA = {}
for bnum, blabel in zip(BATCH_NUMS, BATCHES):
    d = load_batch(bnum)
    if d:
        DATA[blabel] = d
        print(f"  {blabel}: orig={list(d['orig'].keys())[:4]}... "
              f"  nv_orig keys={list(d['nv_orig'].keys())[:4]}...")
    else:
        DATA[blabel] = None
        print(f"  {blabel}: MISSING")

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def savefig(name, fig=None):
    path = os.path.join(OUT, f'{name}.png')
    (fig or plt).tight_layout()
    (fig or plt).savefig(path, dpi=150, bbox_inches='tight')
    plt.close('all')
    print(f"  ✓ {name}.png")

def get_val(batch, mode, vuln, tool):
    """Safely get efficiency value; returns 0 for None/missing."""
    if DATA[batch] is None:
        return 0
    d = DATA[batch][mode]
    return d.get(vuln, {}).get(tool) or 0

def get_overall(batch, mode, tool):
    if DATA[batch] is None:
        return 0
    tc = DATA[batch].get('tool_comparison', {})
    if tool in tc and tc[tool].get(mode) is not None:
        return tc[tool][mode]
    # fallback: read from OVERALL row in eff sheet
    d = DATA[batch][mode]
    return d.get('OVERALL', {}).get(tool) or 0

# ─────────────────────────────────────────────────────────────
# PLOT 1 — Overall efficiency: Original vs Obfuscated (all batches)
# ─────────────────────────────────────────────────────────────
print("\nGenerating plots...")
print("  Plot 1: Overall efficiency (all batches)...")
fig, axes = plt.subplots(1, 5, figsize=(22, 6), sharey=True)
fig.suptitle(
    'Tool Overall Detection Efficiency: Original vs Obfuscated Contracts — Batches 01–05\n'
    r'$E(t) = \sum D(t,v) / \sum N(v) \times 100$  (prevalence-weighted)',
    fontsize=13, fontweight='bold', y=1.03)

x  = np.arange(len(TOOLS))
w  = 0.35

for bi, (batch, ax) in enumerate(zip(BATCHES, axes)):
    ov = [get_overall(batch, 'orig', t) for t in TOOLS]
    ob = [get_overall(batch, 'obf',  t) for t in TOOLS]

    bars_o = ax.bar(x - w/2, ov, w, color=[COLORS[t] for t in TOOLS],
                    alpha=0.88, edgecolor='white', label='Original')
    bars_b = ax.bar(x + w/2, ob, w, color=[COLORS[t] for t in TOOLS],
                    alpha=0.42, edgecolor=[COLORS[t] for t in TOOLS],
                    linewidth=1.3, hatch='//', label='Obfuscated')

    for bar in bars_o:
        h = bar.get_height()
        if h > 1:
            ax.text(bar.get_x() + bar.get_width()/2, h + 1.5,
                    f'{h:.0f}%', ha='center', va='bottom', fontsize=7, fontweight='bold')
    for bar in bars_b:
        h = bar.get_height()
        if h > 1:
            ax.text(bar.get_x() + bar.get_width()/2, h + 1.5,
                    f'{h:.0f}%', ha='center', va='bottom', fontsize=7)

    ax.set_title(f'Batch {batch[1:]}', fontsize=12, fontweight='bold', color=BCOLORS[bi])
    ax.set_xticks(x)
    ax.set_xticklabels(TOOLS, fontsize=8, rotation=22, ha='right')
    ax.set_ylim(0, 120)
    ax.set_ylabel('Efficiency (%)' if bi == 0 else '', fontsize=9)
    ax.axhline(50, color='grey', ls='--', lw=0.7, alpha=0.5)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

legend_el = [
    mpatches.Patch(facecolor='#555', label='Original (solid)'),
    mpatches.Patch(facecolor='#aaa', hatch='//', label='Obfuscated (hatched)'),
]
fig.legend(handles=legend_el, loc='lower center', ncol=2, fontsize=10,
           bbox_to_anchor=(0.5, -0.07))
savefig('01_overall_efficiency_all_batches', fig)

# ─────────────────────────────────────────────────────────────
# PLOT 2 & 3 — Per-vulnerability efficiency per batch (orig + obf)
# ─────────────────────────────────────────────────────────────
for mode, label, code in [('orig', 'Original', '02'), ('obf', 'Obfuscated', '03')]:
    print(f"  Plot {code}: Per-vulnerability ({label}) per batch...")
    for bi, batch in enumerate(BATCHES):
        fig, ax = plt.subplots(figsize=(14, 5))
        x   = np.arange(len(VULNS_PLOT))
        n   = len(TOOLS)
        w   = 0.18
        offs = np.linspace(-(n-1)*w/2, (n-1)*w/2, n)

        any_data = False
        for ti, t in enumerate(TOOLS):
            vals = [get_val(batch, mode, v, t) for v in VULNS_PLOT]
            if sum(vals) > 0:
                any_data = True
            kw = dict(label=t, color=COLORS[t], edgecolor='white', alpha=0.85)
            if mode == 'obf':
                kw.update(hatch='//', alpha=0.75, edgecolor=COLORS[t])
            bars = ax.bar(x + offs[ti], vals, w, **kw)
            for bar, v in zip(bars, vals):
                if v > 3:
                    ax.text(bar.get_x() + bar.get_width()/2,
                            bar.get_height() + 1.5,
                            f'{v:.0f}', ha='center', va='bottom',
                            fontsize=6.5, fontweight='bold')

        ax.set_title(
            f'Batch {batch[1:]} — {label} Contracts: Tool Detection Efficiency per Vulnerability\n'
            r'$E(t,v) = D(t,v) / N(v) \times 100$',
            fontsize=12, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(VULNS_PLOT, fontsize=9, rotation=20, ha='right')
        ax.set_ylim(0, 118)
        ax.set_ylabel('Efficiency (%)', fontsize=10)
        ax.axhline(50, color='grey', ls='--', lw=0.7, alpha=0.5)
        ax.axhline(75, color='steelblue', ls=':', lw=0.7, alpha=0.5)
        ax.legend(loc='upper right', fontsize=9)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        # Annotation for B03 Mythril config issue
        if batch == 'B03' and mode == 'orig':
            ax.text(0.01, 0.97,
                    '* Mythril 0% on originals = Solc compiler-config issue, NOT a detection failure',
                    transform=ax.transAxes, fontsize=8, color='red',
                    style='italic', va='top')

        # Collapse watermark
        if mode == 'obf' and not any_data:
            ax.text(0.5, 0.52, 'COMPLETE DETECTION COLLAPSE\n100% Solc Compilation Failure',
                    transform=ax.transAxes, fontsize=15, color='#C00000',
                    fontweight='bold', ha='center', va='center', alpha=0.30)
        elif mode == 'obf' and batch == 'B02':
            ax.text(0.5, 0.52,
                    'NEAR-TOTAL COLLAPSE\n98.6% Solc Failure\n(1 detection retained: Slither — Frozen Ether)',
                    transform=ax.transAxes, fontsize=11, color='#C00000',
                    fontweight='bold', ha='center', va='center', alpha=0.30)

        savefig(f'{code}_per_vuln_{mode}_{batch}', fig)

# ─────────────────────────────────────────────────────────────
# PLOT 4 — Overall efficiency trend across batches per tool
# ─────────────────────────────────────────────────────────────
print("  Plot 4: Efficiency trend across batches...")
fig, axes = plt.subplots(2, 2, figsize=(16, 11))
fig.suptitle(
    'Tool Overall Efficiency Trend: Original → Obfuscated Across Batches 01–05\n'
    'Solid line = Original | Dashed line = Obfuscated',
    fontsize=13, fontweight='bold')

for ax, t in zip(axes.flat, TOOLS):
    orig_line = [get_overall(b, 'orig', t) for b in BATCHES]
    obf_line  = [get_overall(b, 'obf',  t) for b in BATCHES]
    x_idx = range(len(BATCHES))

    ax.plot(BATCHES, orig_line, 'o-', color=COLORS[t], lw=2.5, ms=8,
            label='Original', zorder=3)
    ax.plot(BATCHES, obf_line, 's--', color=COLORS[t], lw=2.0, ms=7,
            alpha=0.65, label='Obfuscated', zorder=3)

    for i, (o, b) in enumerate(zip(orig_line, obf_line)):
        ax.annotate(f'{o:.0f}%', (BATCHES[i], o),
                    textcoords='offset points', xytext=(0, 9),
                    ha='center', fontsize=8, color=COLORS[t], fontweight='bold')
        if b > 0:
            ax.annotate(f'{b:.0f}%', (BATCHES[i], b),
                        textcoords='offset points', xytext=(0, -14),
                        ha='center', fontsize=8, color=COLORS[t], alpha=0.8)

    ax.fill_between(list(x_idx), orig_line, obf_line,
                    alpha=0.07, color=COLORS[t])

    ax.set_title(t, fontsize=12, fontweight='bold', color=COLORS[t])
    ax.set_xticks(list(x_idx))
    ax.set_xticklabels(BATCHES, fontsize=10)
    ax.set_ylim(0, 118)
    ax.set_ylabel('Overall Efficiency (%)', fontsize=9)
    ax.axhline(50, color='grey', ls='--', lw=0.6, alpha=0.4)
    ax.legend(fontsize=9, loc='upper right')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.set_facecolor('#FAFAFA')
    if t == 'Mythril':
        ax.annotate('*config\nissue', xy=('B03', 0), xytext=(2, 15),
                    textcoords='offset points', fontsize=8,
                    color='red', style='italic', ha='center')

savefig('04_tool_trend_orig_vs_obf', fig)

# ─────────────────────────────────────────────────────────────
# PLOT 5 & 6 — Heatmap: tool × vulnerability (orig and obf avg)
# ─────────────────────────────────────────────────────────────
for mode, label, code in [('orig', 'Original', '05'), ('obf', 'Obfuscated', '06')]:
    print(f"  Plot {code}: Heatmap tool×vuln ({label})...")
    matrix = []
    for t in TOOLS:
        row = []
        for v in VULNS_PLOT:
            vals = [get_val(b, mode, v, t) for b in BATCHES
                    if DATA[b] is not None and
                    DATA[b][mode].get(v, {}).get(t) is not None]
            row.append(np.mean(vals) if vals else 0.0)
        matrix.append(row)
    matrix = np.array(matrix)

    fig, ax = plt.subplots(figsize=(13, 5))
    im = ax.imshow(matrix, cmap='RdYlGn', aspect='auto', vmin=0, vmax=100)
    plt.colorbar(im, ax=ax, label='Efficiency (%)', shrink=0.8)
    ax.set_xticks(range(len(VULNS_PLOT)))
    ax.set_xticklabels(VULNS_PLOT, fontsize=10, rotation=20, ha='right')
    ax.set_yticks(range(len(TOOLS)))
    ax.set_yticklabels(TOOLS, fontsize=11, fontweight='bold')
    sub = 'N/A cells excluded from avg; 0 = tool detected nothing' if mode == 'orig' \
          else 'N/A (vuln absent) excluded; most 0s due to Solc failure'
    ax.set_title(
        f'Average Tool Efficiency per Vulnerability — {label} Contracts (B01–B05 averaged)\n{sub}',
        fontsize=11, fontweight='bold')
    for i in range(len(TOOLS)):
        for j in range(len(VULNS_PLOT)):
            val = matrix[i, j]
            color = 'white' if val > 65 or val < 15 else 'black'
            ax.text(j, i, f'{val:.0f}%', ha='center', va='center',
                    fontsize=10, fontweight='bold', color=color)
    savefig(f'{code}_heatmap_{mode}_avg', fig)

# ─────────────────────────────────────────────────────────────
# PLOT 7 — Radar chart: per-vulnerability per batch (original)
# ─────────────────────────────────────────────────────────────
print("  Plot 7: Radar charts (original)...")
N      = len(VULNS_PLOT)
angles = np.linspace(0, 2*np.pi, N, endpoint=False).tolist()
angles += angles[:1]

fig, axes = plt.subplots(1, 5, figsize=(24, 5), subplot_kw=dict(polar=True))
fig.suptitle(
    'Tool Detection Efficiency per Vulnerability — Original Contracts\n'
    'Radar Chart per Batch (B01–B05)',
    fontsize=13, fontweight='bold', y=1.04)

for bi, (batch, ax) in enumerate(zip(BATCHES, axes)):
    ax.set_title(f'Batch {batch[1:]}', size=12, fontweight='bold',
                 color=BCOLORS[bi], pad=14)
    for t in TOOLS:
        vals = [get_val(batch, 'orig', v, t) for v in VULNS_PLOT]
        vals += vals[:1]
        ax.plot(angles, vals, 'o-', lw=1.8, color=COLORS[t], label=t, ms=4)
        ax.fill(angles, vals, alpha=0.07, color=COLORS[t])
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels([v.replace(' ', '\n') for v in VULNS_PLOT], fontsize=7.5)
    ax.set_ylim(0, 100)
    ax.set_yticks([25, 50, 75, 100])
    ax.set_yticklabels(['25', '50', '75', '100'], fontsize=6.5)
    ax.grid(color='grey', ls='--', lw=0.5, alpha=0.5)

handles = [mpatches.Patch(color=COLORS[t], label=t) for t in TOOLS]
fig.legend(handles=handles, loc='lower center', ncol=4, fontsize=10,
           bbox_to_anchor=(0.5, -0.09))
savefig('07_radar_orig_per_batch', fig)

# ─────────────────────────────────────────────────────────────
# PLOT 8 — N(v) consensus positives stacked bar
# ─────────────────────────────────────────────────────────────
print("  Plot 8: N(v) consensus positives stacked bar...")
fig, ax = plt.subplots(figsize=(14, 6))
x       = np.arange(len(VULNS_PLOT))
bottoms = np.zeros(len(VULNS_PLOT))

for bi, batch in enumerate(BATCHES):
    vals = []
    for v in VULNS_PLOT:
        nv = 0
        if DATA[batch] is not None:
            nv = DATA[batch]['nv_orig'].get(v, 0) or 0
        vals.append(nv)
    bars = ax.bar(x, vals, bottom=bottoms, label=f'Batch {batch[1:]}',
                  color=BCOLORS[bi], alpha=0.82, edgecolor='white')
    for bar, val, bot in zip(bars, vals, bottoms):
        if val > 5:
            ax.text(bar.get_x() + bar.get_width()/2, bot + val/2,
                    str(val), ha='center', va='center',
                    fontsize=8, color='white', fontweight='bold')
    bottoms += np.array(vals, dtype=float)

ax.set_xticks(x)
ax.set_xticklabels(VULNS_PLOT, fontsize=10, rotation=20, ha='right')
ax.set_ylabel('N(v) — Consensus Positives (contracts)', fontsize=10)
ax.set_title(
    'N(v) Consensus Positives per Vulnerability — Original Contracts (B01–B05 stacked)\n'
    'N(v) = contracts where ANY tool detected the vulnerability (ground truth)',
    fontsize=11, fontweight='bold')
ax.legend(title='Batch', fontsize=9, title_fontsize=9)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
savefig('08_nv_positives_stacked', fig)

# ─────────────────────────────────────────────────────────────
# PLOT 9 — Efficiency drop heatmap: tool × batch
# ─────────────────────────────────────────────────────────────
print("  Plot 9: Efficiency drop heatmap (orig→obf per tool×batch)...")
drop_matrix = []
for t in TOOLS:
    row = []
    for batch in BATCHES:
        if DATA[batch] is None:
            row.append(np.nan)
            continue
        tc = DATA[batch].get('tool_comparison', {}).get(t, {})
        drop = tc.get('drop')
        if drop is None:
            o = get_overall(batch, 'orig', t)
            b = get_overall(batch, 'obf',  t)
            drop = o - b if (o + b) > 0 else np.nan
        row.append(float(drop) if drop is not None else np.nan)
    drop_matrix.append(row)
drop_matrix = np.array(drop_matrix, dtype=float)

# Diverging colormap: red = big drop, blue = gain
vmax = np.nanmax(np.abs(drop_matrix)) if not np.all(np.isnan(drop_matrix)) else 60
fig, ax = plt.subplots(figsize=(11, 5))
cmap = plt.get_cmap('RdBu_r')
cmap.set_bad(color='#DDDDDD')
im = ax.imshow(drop_matrix, cmap=cmap, aspect='auto',
               vmin=-vmax, vmax=vmax)
plt.colorbar(im, ax=ax, label='Efficiency Drop (pp)', shrink=0.8)
ax.set_xticks(range(len(BATCHES)))
ax.set_xticklabels([f'Batch {b[1:]}' for b in BATCHES], fontsize=10)
ax.set_yticks(range(len(TOOLS)))
ax.set_yticklabels(TOOLS, fontsize=11, fontweight='bold')
ax.set_title(
    'Obfuscation Impact: Efficiency Drop (Orig − Obf) per Tool × Batch\n'
    'Red = large loss | Blue = gain | Grey = N/A (Solc failure)',
    fontsize=11, fontweight='bold')
for i in range(len(TOOLS)):
    for j in range(len(BATCHES)):
        val = drop_matrix[i, j]
        if not np.isnan(val):
            color = 'white' if abs(val) > vmax * 0.6 else 'black'
            ax.text(j, i, f'{val:+.0f}pp', ha='center', va='center',
                    fontsize=10, fontweight='bold', color=color)
        else:
            ax.text(j, i, 'N/A', ha='center', va='center',
                    fontsize=9, color='#888888', style='italic')
savefig('09_efficiency_drop_heatmap', fig)

# ─────────────────────────────────────────────────────────────
# PLOT 10 — Robustness to obfuscation grouped bar
# ─────────────────────────────────────────────────────────────
print("  Plot 10: Robustness to obfuscation bar chart...")
fig, ax = plt.subplots(figsize=(13, 6))
x   = np.arange(len(BATCHES))
w   = 0.18
offs = np.linspace(-(len(TOOLS)-1)*w/2, (len(TOOLS)-1)*w/2, len(TOOLS))

for ti, t in enumerate(TOOLS):
    vals = []
    for batch in BATCHES:
        if DATA[batch] is None:
            vals.append(0)
            continue
        tc  = DATA[batch].get('tool_comparison', {}).get(t, {})
        rob = tc.get('robustness')
        if rob is None:
            o = get_overall(batch, 'orig', t)
            b = get_overall(batch, 'obf',  t)
            rob = (b / o * 100) if o > 0 else 0
        vals.append(min(float(rob) if rob else 0, 200))  # cap at 200% for display
    bars = ax.bar(x + offs[ti], vals, w, label=t,
                  color=COLORS[t], alpha=0.85, edgecolor='white')
    for bar, v in zip(bars, vals):
        if v > 3:
            ax.text(bar.get_x() + bar.get_width()/2,
                    bar.get_height() + 1.5,
                    f'{v:.0f}%', ha='center', va='bottom', fontsize=7, fontweight='bold')

ax.axhline(100, color='black', ls='-', lw=1.2, alpha=0.4, label='100% = no loss')
ax.axhline(50, color='grey', ls='--', lw=0.8, alpha=0.4)
ax.set_xticks(x)
ax.set_xticklabels([f'Batch {b[1:]}' for b in BATCHES], fontsize=10)
ax.set_ylim(0, 220)
ax.set_ylabel('Robustness Score R(t) = E_obf / E_orig × 100 (%)', fontsize=9)
ax.set_title(
    'Tool Robustness to Obfuscation: R(t) = E_obf(t) / E_orig(t) × 100\n'
    '100% = no loss | >100% = improved (small N) | 0% = complete failure',
    fontsize=11, fontweight='bold')
ax.legend(fontsize=9)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
savefig('10_tool_robustness_bar', fig)

# ─────────────────────────────────────────────────────────────
# SUMMARY
# ─────────────────────────────────────────────────────────────
print(f"\n{'='*58}")
print(f"  All plots saved to: {OUT}")
files = sorted(f for f in os.listdir(OUT) if f.endswith('.png'))
total_kb = 0
for f in files:
    kb = os.path.getsize(os.path.join(OUT, f)) // 1024
    total_kb += kb
    print(f"  {f:52s}  {kb:4d} KB")
print(f"{'='*58}")
print(f"  Total: {len(files)} PNGs  ({total_kb} KB)")
